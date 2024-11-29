import os

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import openpyxl
import time
import zipfile
import datetime
from django.conf import settings
from mys.templateAnswerClassFile import templateAnswerClass
from mys.WorkbookProcessor import Processor
from mys.getAnswerArrays import flattenedAnswers
import shutil
from mys.monitor import Monitor
import mys.marking  
from mys.consumers import ChatConsumer


def progPrint(*args):
    textString =""
    for x in args:
        textString = textString + x
    print("Textstring is: ", textString)
    localConsumer = ChatConsumer.getChannelToBrowser()
    localConsumer.sendMessage(textString)

class setupMain():
    def __init__(self, template_file, studentFiles, request, option ):
        
        # Allowed file types
        self.filetypes = [('All files', '.*'), ("Excel files", ".xl*")]
        # Add empty references to template and folder location - will be filled by functions
        self.template = template_file
        self.folder = studentFiles
        self.option = option
        self.request = request
        #Setup all the elements in the mainform.
    
    
    # a bit hacky, but adds the file names for now to the text window with a message
    # will obviously need to be re-done for actual processing
    def process(self):
        # prevent button doing anything if empty string
        if not self.folder == "":
            # quite complicated, but is a list of student files in the folder
            files = (file for file in os.listdir(self.folder) if os.path.isfile(os.path.join(self.folder, file)))
            FileMonitor = Monitor()
            # reduce template path to filename
            template_file = self.template
            # load workbooks for feeding in below
            DO_template_wb=openpyxl.load_workbook(template_file, data_only=True)
            template_wb=openpyxl.load_workbook(template_file)
            # start time for processing
            start_time = time.time()
            # name of output folder
            output_folder = "Output"
            # join to create new path name
            next_folder = os.path.join(self.folder , output_folder)
            # move to student file directory
            os.chdir(self.folder)
            if not os.path.exists(output_folder):
                os.mkdir(output_folder)
            elif os.path.exists(output_folder):
                try:
                    shutil.rmtree(output_folder)
                    os.mkdir(output_folder)
                except PermissionError:
                    print("Please close output folder and any Excel files.")
                    progPrint("Please close output folder and any Excel files.")
            # go through all files
            
            flattenedTemplates = flattenedAnswers(template_wb, self.option)
            #WorkbookManager = Processor(DO_template_wb, template_wb, next_folder)
            WorkbookManager = Processor(DO_template_wb, flattenedTemplates, next_folder)
            for student_file in files:
                # change directory to the chosen folder so files can be used
                # this is redundant first time around, but it needs to change back next iteration
                os.chdir(self.folder)
                # only work with them in xl* files
                if ".xl" in student_file:
                    print("end", "Processing "+ student_file + "\n")
                    progPrint("end", "Processing "+ student_file + "\n")
                   
                    # begin deconstruction of workbooks, then sheets, then comparisons
                    WorkbookManager.process_workbooks(student_file)
                    #self.process_workbooks(DO_template_wb, template_wb, elements, student_file, next_folder)
                    #self.parent.update_idletasks()
            # print out time taken
            elapsed_time = time.time() - start_time
            print( "Processing required {} seconds".format(elapsed_time))
            progPrint( "Processing required {} seconds".format(elapsed_time))
            # show folder to user
            #os.startfile(next_folder)
            liveLinkPath = os.path.join(self.folder ,"Output")
            # LiveLink marking will fail if output folder is empty (all files rejected)
            # output_empty returns a boolean after checking output folder
            
            output_empty = FileMonitor.output_folder_empty(liveLinkPath)
            if output_empty:
                print("There are problems with all files - processing aborted. Please check output window for errors.")
                progPrint("There are problems with all files - processing aborted. Please check output window for errors.")
                
                self.submit.config(state="normal")
            
            else:
                print("Live linking output files to summary sheet.\n")
                progPrint("Live linking output files to summary sheet.\n")
                mys.marking.liveLink(liveLinkPath)
            #os.startfile(next_folder)
            progPrint("Preparing to zip up marked results.\n")        
            #    
            #zip file
            user_dir_path_name = os.path.join(settings.MEDIA_ROOT,str(self.request.user.uuid))
            results_folder = os.path.join(user_dir_path_name , "results.zip")
            if os.path.exists(results_folder):
               os.remove(results_folder)
               print("Previous output file removed")
            zf = zipfile.ZipFile(results_folder, "w")
            for dirname, subdirs, files in os.walk(next_folder):
                zf.write(dirname)
                for filename in files:
                    zf.write(os.path.join(dirname, filename))
            zf.close()
            progPrint("Results have been written to zip archive.\n")
            os.chdir(user_dir_path_name)
            progPrint("***ENABLE_BUTTON***")        
 


        
