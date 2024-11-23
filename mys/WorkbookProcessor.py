import os
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Alignment,  Side, Fill
from openpyxl.styles import colors
import openpyxl
import time
import datetime
import shutil


class Processor:
    def __init__(self, DO_template, flattened_sheets, save_dir):
        # construct template wb, dict of flattened sheets, where to save files, and DO sheet names
        self.DO_template_wb= DO_template
        self.flattened_sheets = flattened_sheets.workbook_dict
        self.save_dir = save_dir
        self.DO_template_sheets = self.DO_template_wb.sheetnames
        self.Student_Named_File = ""


    # work through comparisons of each template and student sheet (4 sheets as parameters)
    # doesn't matter which sheets, as all same length
    def process_workbooks(self, student_workbook):
        """Gets  workbooks and their sheets ready for processing, then passes them to compare function"""
        # set up student workbooks and save their sheet names as a list
        DO_student_wb = openpyxl.load_workbook(student_workbook, data_only=True)
        student_wb = openpyxl.load_workbook(student_workbook)
        DO_student_sheets = DO_student_wb.sheetnames
        student_sheets = student_wb.sheetnames
        # iterate through sheets
        
        #for sheet in range(len(self.DO_template_sheets)):
        for template_sheet in self.flattened_sheets:
            # Get name of flattened template sheet
            template_sheet_name = self.flattened_sheets[template_sheet]["sheet_name"]
            template_sheet_object_array = self.flattened_sheets[template_sheet]["template_cell_array"]
            # essentially all sheets of a workbook one at a time
            # list taken from dictionary - might be a small speed increase
            # all others are sheet names
            DO_template_sheet = self.DO_template_wb[template_sheet_name]
            DO_student_sheet = DO_student_wb[template_sheet_name]
            student_sheet = student_wb[template_sheet_name]
            # compare sheets
            self.compare_sheets(DO_template_sheet, template_sheet_object_array,
            DO_student_sheet, student_sheet, student_wb)
        # finally, change directory and save changes made to student sheet once student workbook done
        os.chdir(self.save_dir)
        student_wb.save(self.Student_Named_File)

    def compare_sheets(self, DO_template_sheet, sheet_from_dict, DO_student_sheet, student_sheet, student_wb):
        """Compares each individual sheet against others"""
        # go through the cells of the flattened template array in dict
        counts_correct = 0
        counts_formulas = 0
        counts_incorrect = 0
        
        for i in range(len(sheet_from_dict)):
                # coordinates taken from templateAnswerClass objects
                coordinateX= sheet_from_dict[i].cellRefTuple[0]
                coordinateY= sheet_from_dict[i].cellRefTuple[1]
                templ_cell= sheet_from_dict[i].value
                DO_templ_cell= DO_template_sheet.cell(coordinateX, coordinateY).value
                # student cells
                DO_student_cell = DO_student_sheet.cell(coordinateX, coordinateY).value
                student_cell = student_sheet.cell(coordinateX, coordinateY).value
                # Now do comparisons:
                # ignore all cells that are empty or date
                if not (DO_templ_cell == None and DO_student_cell == None) and \
                   not (type(DO_templ_cell) == datetime.datetime and type(DO_student_cell) == datetime.datetime):
                    # case 1: both are numbers
                    if not (isinstance(DO_templ_cell, str) and isinstance(DO_student_cell, str)):
                        # numbers match
                        if DO_templ_cell == DO_student_cell:
                        # mark green
                            student_sheet.cell(coordinateX, coordinateY).fill=PatternFill("solid", fgColor="33FF8A")
                            counts_correct +=1
                        # numbers don't match but formulas do
                        elif DO_templ_cell != DO_student_cell and templ_cell == student_cell:
                        # mark yellow
                            student_sheet.cell(coordinateX, coordinateY).fill=PatternFill("solid", fgColor="FEF561")
                            counts_formulas +=1
                        # neither numbers nor formulas match
                        elif DO_templ_cell != DO_student_cell and templ_cell != student_cell:
                        # mark yellow
                            student_sheet.cell(coordinateX, coordinateY).fill=PatternFill("solid", fgColor="FF7f7F")
                            counts_incorrect +=1

         # After the highlighting of compared cells has occurred the following code outputs the 
        # count results (correct, formulas, and incorrect) to the details sheets of the marked student
        # worksheet. A small amount of formatting occurs with the outputing of the data.                    
        student_sheet_details = student_wb.get_sheet_by_name("Details")
        student_sheet_details.merge_cells('E17:F17')
        student_sheet_details.merge_cells('H17:I17')
        student_sheet_details.merge_cells('K17:L17')
        student_sheet_details['E16'].font = Font(name='arial', size =14)
        student_sheet_details['H16'].font = Font(name='arial', size =14)
        student_sheet_details['K16'].font = Font(name='arial', size =14)
        student_sheet_details['E17'].font = Font(name='arial', size =14)
        student_sheet_details['H17'].font = Font(name='arial', size =14)
        student_sheet_details['K17'].font = Font(name='arial', size =14)
        student_sheet_details['E18'].font = Font(name='arial', size =14)
        student_sheet_details['H18'].font = Font(name='arial', size =14)
        student_sheet_details['K18'].font = Font(name='arial', size =14)
        student_sheet_details["E16"].value = "Correct Answer"
        student_sheet_details["H16"].value = "Formulas Correct"
        student_sheet_details["K16"].value = "Incorrect Answer"
        student_sheet_details['E17'].alignment = Alignment(horizontal='center')
        student_sheet_details['H17'].alignment = Alignment(horizontal='center')
        student_sheet_details['K17'].alignment = Alignment(horizontal='center')
        student_sheet_details['E18'].alignment = Alignment(horizontal='center')
        student_sheet_details['H18'].alignment = Alignment(horizontal='center')
        student_sheet_details['K18'].alignment = Alignment(horizontal='center')
        student_sheet_details["E17"].value = counts_correct
        student_sheet_details["H17"].value = counts_formulas
        student_sheet_details["K17"].value = counts_incorrect
        student_sheet_details["E18"].value = "=ROUND(E17/(E17+H17+K17)*100,1)"
        student_sheet_details["H18"].value = "=ROUND(H17/(E17+H17+K17)*100,1)"
        student_sheet_details["K18"].value = "=ROUND(K17/(E17+H17+K17)*100,1)"
        student_sheet_details["C17"].value = "RAW"
        student_sheet_details["C18"].value = "%"
        student_sheet_details["C19"].value = "AUTOMARK"
        student_sheet_details.merge_cells("E19:F19")
        student_sheet_details["E19"].alignment = Alignment(horizontal='center')
        student_sheet_details['E19'].font = Font(name='arial', size =14)
        student_sheet_details["E19"].value = "=ROUND(E18/100*F8,0)"
        student_sheet_details['C17'].font = Font(name='arial', size =14)
        student_sheet_details['C18'].font = Font(name='arial', size =14)
        student_sheet_details['C19'].font = Font(name='arial', size =14)

        student_sheet_details.merge_cells('E18:F18')
        student_sheet_details.merge_cells('H18:I18')
        student_sheet_details.merge_cells('K18:L18')

        
        self.set_range_color(student_sheet_details, "E16:G19", PatternFill("solid", fgColor="33FF8A"))
        self.set_range_color(student_sheet_details, "H16:J18", PatternFill("solid", fgColor="FEF561"))
        self.set_range_color(student_sheet_details, "K16:M18", PatternFill("solid", fgColor="FF7f7F"))
           
        #student_sheet_details['D18'].font = Font(name='arial', size =14)
        #student_sheet_details['G18'].font = Font(name='arial', size =14)
        #student_sheet_details['J18'].font = Font(name='arial', size =14)
        #student_sheet_details["N8"].value = "Correct Answer"
        #student_sheet_details["Q8"].value = "Formulas Correct"
        #student_sheet_details["T8"].value = "Incorrect Answer"
        #student_sheet_details['D18'].alignment = Alignment(horizontal='center')
        #student_sheet_details['G18'].alignment = Alignment(horizontal='center')
        #student_sheet_details['J18'].alignment = Alignment(horizontal='center')
        #student_sheet_details["D18"].value = "0"
        #student_sheet_details["G18"].value = "0"
        #student_sheet_details["J18"].value = "0"

        student_sheet_details["D14"].value = "Automated Marking"
        student_sheet_details['D14'].font = Font(name='arial', size =14)
        self.set_border(student_sheet_details, "C13:M20")

        student_sheet_details["C23"].value = "Lecturer's Mark"
        student_sheet_details['C23'].font = Font(name='arial', size =14)
        student_sheet_details.merge_cells('C24:D24')
        student_sheet_details['C24'].font = Font(name='arial', size =14)
        student_sheet_details['C24'].value="0"
        student_sheet_details['C24'].fill =  PatternFill("solid", fgColor="33FF8A")

        # Create Output Marked Workbook name
        self.Student_Named_File = student_sheet_details["F11"].value + " " + student_sheet_details["C11"].value + " " + str(student_sheet_details["I11"].value) + ".xlsx"
        



    def set_range_color(self, ws, cell_range, fill_colour):
        rows = ws[cell_range]
        #fill = fill_colour
        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                cell.fill = fill_colour

    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='thick', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border