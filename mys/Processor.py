import os
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import openpyxl
import time
import datetime
#from templateAnswerClassFile import templateAnswerClass
import shutil


class Processor():
        def __init__(self, DO_template, template, save_dir):
		    # construct parameters and initialise others for load method to save complexity
            self.save_dir = save_dir
            self.DO_template_sheets = DO_template.sheetnames
            self.template_sheets = template.sheetnames
            self.student_workbook =""

        # work through comparisons of each template and student sheet (4 sheets as parameters)
        # doesn't matter which sheets, as all same length
        def process_workbooks(self, student_workbook):
            # set up student workbooks and save their sheet names as a list
            self.student_workbook = student_workbook
            DO_student_wb = openpyxl.load_workbook(self.student_workbook, data_only=True)
            student_wb = openpyxl.load_workbook(self.student_workbook)
            DO_student_sheets = DO_student_wb.sheetnames
            student_sheets = student_wb.sheetnames
            # iterate through sheets
            for sheet in range(len(self.DO_template_sheets)):
                # For understanding: one sheet below = a workbook[list of its sheets [whichever sheet as an index]]
                # essentially all sheets of a workbook one at a time
                DO_template_sheet = self.DO_template[self.DO_template_sheets[sheet]]
                template_sheet = self.template[self.template_sheets[sheet]]
                DO_student_sheet = DO_student_wb[self.DO_student_sheets[sheet]]
                student_sheet = student_wb[self.student_sheets[sheet]]
                # compare shees
                self.compare_sheets(DO_template_sheet, template_sheet,
                                   DO_student_sheet, student_sheet)
                # change directory and save changes made to student sheet
                os.chdir(self.save_dir)
                self.student_wb.save(self.workbook)

    # compare each individual sheet
        def compare_sheets(self, DO_template_sheet, template_sheet, DO_student_sheet, student_sheet):
            # go through the cells of the template sheet and compare to student sheet
                for i in range(1, DO_template_sheet.max_row+1):
                    for j in range(1, DO_template_sheet.max_column+1):
                        DO_templ_cell= DO_template_sheet.cell(i, j).value
                        templ_cell= template_sheet.cell(i, j).value
                        # student cells
                        DO_student_cell = DO_student_sheet.cell(i, j).value
                        student_cell = student_sheet.cell(i, j).value
                    # Now do comparisons:
                        # ignore all cells that are empty or date
                        if not (DO_templ_cell == None and DO_student_cell == None) and \
                        not (type(DO_templ_cell) == datetime.datetime and type(DO_student_cell) == datetime.datetime):
                          # case 1: both are numbers
                            if not (isinstance(DO_templ_cell, str) and isinstance(DO_student_cell, str)):
                                # numbers match
                                if DO_templ_cell == DO_student_cell:
                                    # mark green
                                    student_sheet.cell(i, j).fill=PatternFill("solid", fgColor="33FF8A")
                                # numbers don't match but formulas do
                                elif DO_templ_cell != DO_student_cell and templ_cell == student_cell:
                                    # mark yellow
                                    student_sheet.cell(i, j).fill=PatternFill("solid", fgColor="FEF561")
                                # neither numbers nor formulas match
                                elif DO_templ_cell != DO_student_cell and templ_cell != student_cell:
                                    # mark yellow
                                    student_sheet.cell(i, j).fill=PatternFill("solid", fgColor="FF7f7F")


