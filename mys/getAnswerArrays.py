import openpyxl
from mys.templateAnswerClassFile import templateAnswerClass

class flattenedAnswers:
    def __init__(self, template, option):
        # option assignment here
        self.default = False
        self.highlighted = False
        self.named = False
        self.choose_option(option)
        self.template = template
        # this is a dictionary of flattened arrays, indexed by sheet name
        self.workbook_dict = {}
        self.flattenWorkBooks()


    def choose_option(self, option):
        if option == "all":
            self.default = True
        elif option == "highlighted":
            self.highlighted = True
        else:
            self.named = True

    def flattenWorkBooks(self):
        """This iterates through the template, depending on chosen option, and creates a flattened dictionary
        for use by workbook processor class"""
        if self.default:
            for template_sheet in self.template.worksheets:
                template_cell_array = []
                for i in range(1, template_sheet.max_row+1):
                    for j in range(1, template_sheet.max_column+1):
                        element = templateAnswerClass((i, j), template_sheet.cell(i, j).value)
                        template_cell_array.append(element)
                # once array complete, index it in dict under sheet name
                worksheet_data={}
                worksheet_data["template_cell_array"]=template_cell_array
                worksheet_data["sheet_name"]=template_sheet.title
                self.workbook_dict[template_sheet.title] = worksheet_data# template_cell_array

        # if highlighting chosen, only difference here is checking for filled cells only
        elif self.highlighted:
            for template_sheet in self.template.worksheets:
                template_cell_array = []
                for i in range(1, template_sheet.max_row+1):
                    for j in range(1, template_sheet.max_column+1):
                        # check for filled cells of any colour
                        if template_sheet.cell(i, j).fill.start_color.index != "00000000":
                            element = templateAnswerClass((i, j), template_sheet.cell(i, j).value)
                            template_cell_array.append(element)
                worksheet_data={}
                worksheet_data["template_cell_array"]=template_cell_array
                worksheet_data["sheet_name"]=template_sheet.title
                self.workbook_dict[template_sheet.title] = worksheet_data# template_cell_array
                #self.workbook_dict[template_sheet.title] = template_cell_array

        # print(self.workbook_dict)
        # defined names - not properly implemented yet as unsure how it best works
      #  else:
      #      template_cell_array = []
      #      rangeNames = [dn for dn in self.template.defined_names.definedName]
		    ##iterate through rangeNames
      #      for template_sheet in self.template.worksheets:
      #          for cellName in rangeNames:
      #              components = cellName.attr_text.split("!") #split the attr_text string into a sheet string and a cell reference string
      #              sheet = wb.get_sheet_by_name(components[0])
      #              print('Range Name: {0}, Actual Range: {1}, and top left range value: {2}.'
      #                    .format(cellName.name, cellName.attr_text, sheet[components[1]].value))
      #              answers = templateAnswerClass(cellName.name, components[0],components[1],sheet[components[1]].value)
      #              template_cell_array.append(answers)
      #          self.workbook_dict[template_sheet.title] = template_cell_array

            
