import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
import os


#walks through and loads each xlsv file.
def liveLink(path):
    os.chdir(path)
    for root, dirs, files in os.walk(".", topdown=True):
        wb1 = xl.Workbook()
        rowCount=3
        greenColour = PatternFill("solid", fgColor="33FF8A")
        yellowColour = PatternFill("solid", fgColor="FEF561")
        redColour = PatternFill("solid", fgColor="FF7F7F")
        cell_alignment = Alignment(horizontal='center')
        for name in files:
           #makes sure dont get an error for reading code program.
          if (name != "Marking.py" and name != "markingSummary.xlsx"):
              if(not name.startswith("~$")):
                  wb = xl.load_workbook(name)
          
                  #loads work sheet with student details on it
                  sheet=wb['Details']
          
                  #sets up live linking to the cells in the workbooks detail sheet
                  #print each students details on new line using rowcount                      
                  ws1 = wb1["Sheet"]
                  ws1["A1"].value = "Last Name"
                  ws1['B1'].value = "First Name"
                  ws1['C1'].value = "Student Id"
                  ws1['C1'].alignment = cell_alignment
                  ws1['D1'].value = "Qualification"
                  ws1['D1'].alignment = cell_alignment
                  ws1['F1'].value = "GreenCount"
                  ws1['F1'].alignment = cell_alignment
                  ws1['G1'].value = "YellowCount"
                  ws1['G1'].alignment = cell_alignment
                  ws1['H1'].value = "RedCount"
                  ws1['H1'].alignment = cell_alignment
                  ws1['J1'].value = "Green%"
                  ws1['J1'].alignment = cell_alignment
                  ws1['K1'].value = "Yellow%"
                  ws1['K1'].alignment = cell_alignment
                  ws1['L1'].value = "Red%"
                  ws1['L1'].alignment = cell_alignment
                  ws1['N1'].value = "Lecturer's Mark"
                  ws1['N1'].alignment = cell_alignment

                  ws1.cell(rowCount,1).value = sheet['F11'].value
                  ws1.cell(rowCount,2).value = sheet['C11'].value
                  ws1.cell(rowCount,3).value = sheet['I11'].value
                  ws1.cell(rowCount,3).alignment = cell_alignment
                  ws1.cell(rowCount,4).value = sheet['L11'].value
                  ws1.cell(rowCount,4).alignment = cell_alignment
                  #ws1.cell(rowCount,5).value = "=A3 & \" \" & B3"
                  ws1.cell(rowCount,6).value = sheet['E17'].value
                  ws1.cell(rowCount,6).fill = greenColour
                  ws1.cell(rowCount,6).alignment = cell_alignment
                  ws1.cell(rowCount,7).value = sheet['H17'].value
                  ws1.cell(rowCount,7).fill = yellowColour
                  ws1.cell(rowCount,7).alignment = cell_alignment

                  ws1.cell(rowCount,8).value = sheet['K17'].value
                  ws1.cell(rowCount,8).fill = redColour
                  ws1.cell(rowCount,8).alignment = cell_alignment
                  #ws1.cell(rowCount,10).value = "='{1}[{0}]Details'!$E$18".format(name,path)
                  #ws1.cell(rowCount,11).value = "='{1}[{0}]Details'!$H$18".format(name,path)
                  #ws1.cell(rowCount,12).value = "='{1}[{0}]Details'!$K$18".format(name,path)
          
                  ws1.cell(rowCount,10).value = "=IF(SUM($F${0}:$H${0})=0,"",ROUND(F{0}/SUM($F${0}:$H${0})*100,1))".format(rowCount)
                  ws1.cell(rowCount,10).fill = greenColour
                  ws1.cell(rowCount,10).alignment = cell_alignment
                  ws1.cell(rowCount,11).value = "=IF(SUM($F${0}:$H${0})=0,"",ROUND(G{0}/SUM($F${0}:$H${0})*100,1))".format(rowCount)
                  ws1.cell(rowCount,11).fill = yellowColour
                  ws1.cell(rowCount,11).alignment = cell_alignment
                  ws1.cell(rowCount,12).value = "=IF(SUM($G${0}:$H${0})=0,"",ROUND(H{0}/SUM($F${0}:$H${0})*100,1))".format(rowCount)
                  ws1.cell(rowCount,12).fill = redColour
                  ws1.cell(rowCount,12).alignment = cell_alignment
                  ws1.cell(rowCount,14).value = sheet['C24'].value
                  ws1.cell(rowCount,14).alignment = cell_alignment

                  #saves over file
                  #moves to next line of students details
                  rowCount += 1
        
                  
        ws1.column_dimensions['A'].width=14
        ws1.column_dimensions['B'].width=14
        ws1.column_dimensions['C'].width=10
        ws1.column_dimensions['D'].width=12
        ws1.column_dimensions['E'].width=8
        ws1.column_dimensions['F'].width=11
        ws1.column_dimensions['G'].width=11
        ws1.column_dimensions['H'].width=10
        ws1.column_dimensions['I'].width=8
        ws1.column_dimensions['J'].width=9
        ws1.column_dimensions['K'].width=9
        ws1.column_dimensions['L'].width=9
        ws1.column_dimensions['M'].width=8
        ws1.column_dimensions['N'].width=16

        wb1.save("markingSummary.xlsx")
        


#liveLink("C:\\Users\\karl bailey\\source\\repos\\MYS\\MYS\\StudentFolder\\Output\\")
